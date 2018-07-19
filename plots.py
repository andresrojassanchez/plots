'''
Natural/Daily Plot Generator Code
Developed by Andres RS of Acumen for Gannett Fleming

This is a simple script that allows the user to select a folder containing
.csv files of data, process them through a pandas dataframe, select a template
based on columns, populate the template with the data, and generate the plots
in Excel.
'''

# Required libaries for pandas/excel/datetime/gui/etc. methods.
import numpy as np
import pandas as pd
import pygsheets
import matplotlib.pyplot as plt
import itertools
import pylab
import os
import datetime as dt
import shutil
import openpyxl
import xlwt
import xlrd
import xlutils
import copy
from openpyxl import load_workbook
from openpyxl import Workbook
import tkFileDialog 
from Tkinter import *
import tkMessageBox

# Function to append/add the data to an existing template.
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, index = False, \
                header = None, **to_excel_kwargs)

    # save the workbook
    writer.save()

# Simple GUI for the user to select the desired directory that contains the
# .csv files of raw data.
root = Tk()
root.withdraw()
dirname = tkFileDialog.askdirectory(parent=root, initialdir="/",
                                    title='Please select a directory')

# Collects all files ending with .csv into a list for processing.
files = [file for file in os.listdir(dirname) if file.endswith(".csv")]
print(files)
# Iterates through the every file found ending with .csv
for file in files:
    # Read the csv data into a pandas dataframe; skipping the first 11 rows
    print(dirname)
    df = pd.read_csv(dirname + '/' + file, skiprows = 11)

    # Change the TIME column to standard datetime format for resampling
    df['TIME'] = pd.to_datetime(df['TIME'])

    # Set the index to TIME, resample using linear interpolation, reset the 
    # index, set the index to sample to move it to the first column, then
    # reset the index one last time.
    df = df.set_index('TIME').resample('1S').interpolate().reset_index() \
                      .set_index('SAMPLE').reset_index()   
    
    # Apply a lambda function to set the datatime back to the original format.
    df['TIME'] = df['TIME'].apply(lambda x: x.strftime('%m/%d/%Y %I:%M:%S %p'))    
    
    # Create the outputfilename by add _natural_plot to the raw csv data
    output_filename = os.path.splitext(file)[0] + '_natural_plot.xlsx'

    # Template selection. If a new template is desired, create another if 
    # state with the number of columns that the template has. Defaults to 9,
    # which is the old .csv format.
    if len(df.columns) == 6:    
    
        # Get the path of the template for the copy operation
        template = os.path.join(os.path.dirname(os.path.realpath(__file__)), \
                                'new_natural_template.xlsx')
        # Make a new copy of the template with the output filename        
        shutil.copy(template, output_filename)
    
        # Add the processed data into the copied excel file
        append_df_to_excel(output_filename, df, sheet_name = 'data', \
                           startrow = 19)
    # Same as above.
    else:
        template = os.path.join(os.path.dirname(os.path.realpath(__file__)), \
                                'old_natural_template.xlsx')
        shutil.copy(template, output_filename)
        append_df_to_excel(output_filename, df, sheet_name = 'data', \
                           startrow = 19)

if plots == 'plots': main()

'''
# GOOGLE SHEETS
# Files too large (cell count > 40K); can't export to Google Sheets. 
gc = pygsheets.authorize(outh_file = '/home/andres/Documents/plots/creds.json')
gc.create('')
sh = gc.open('')    
wks = sh[0]
wks.set_dataframe(df, (1,1))
'''